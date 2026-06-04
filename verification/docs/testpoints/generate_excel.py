#!/usr/bin/env python3
"""
测试点CSV转Excel脚本
V1.0 2026.05.29

将CSV格式的测试点文档转换为带格式的Excel文件(.xlsx)
用法: python generate_excel.py
"""

import csv
import os
import glob

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# 样式定义
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Microsoft YaHei", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 列宽定义
COLUMN_WIDTHS = [30, 30, 35, 40, 40, 40]


def csv_to_xlsx(csv_path, xlsx_path, module_name):
    """将CSV文件转换为格式化的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{module_name}_Testpoints"

    # 读CSV数据
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print(f"Warning: {csv_path} is empty")
        wb.save(xlsx_path)
        return

    # 写标题行
    headers = rows[0]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 设置列宽
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写数据行
    for row_idx, row in enumerate(rows[1:], 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)}"

    wb.save(xlsx_path)
    print(f"Generated: {xlsx_path}")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_files = sorted(glob.glob(os.path.join(script_dir, "*.csv")))

    if not csv_files:
        print("No CSV files found in", script_dir)
        return

    for csv_file in csv_files:
        basename = os.path.splitext(os.path.basename(csv_file))[0]
        # 提取模块名
        module_name = basename.replace("_testpoints", "").upper()
        xlsx_file = os.path.join(script_dir, basename + ".xlsx")
        csv_to_xlsx(csv_file, xlsx_file, module_name)

    print(f"\nDone! {len(csv_files)} Excel files generated.")


if __name__ == "__main__":
    main()
